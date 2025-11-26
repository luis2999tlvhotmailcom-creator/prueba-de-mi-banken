from app.models.deliveries import Delivery, db
from app.models.delivery_detail import DeliveryDetail
from app.models.medications import Medication  # igual que en PURCHASES
from app.models.patients import Patient

from sqlalchemy.exc import IntegrityError
from datetime import datetime

# PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io


# ============================================
# LISTAR SOLO ACTIVOS
# ============================================
def listar_todos():
    return Delivery.query.order_by(Delivery.id).all()


# ============================================
# LISTAR POR ID 
# ============================================
def listar_por_id(id):
    return Delivery.query.filter_by(id=id).first()


# ============================================
# CREAR
# ============================================
def crear(data):
    details_data = data.pop("details", [])

    delivery = Delivery(
        patient_identifier=data.get("patient_identifier"),
        notes=data.get("notes"),
        delivery_date=datetime.now(),
        delivery_state=data.get("delivery_state", "A")
    )
    db.session.add(delivery)
    db.session.flush()  # para obtener delivery.id

    # Calcular total_amount sumando todos los detalles
    total_amount = 0
    for det in details_data:
        amount = det.get("amount", 0) or 0
        detail = DeliveryDetail(
            delivery_identifier=delivery.id,
            medication_identifier=det["medication_identifier"],
            amount=amount,
            observations=det.get("observations", "")
        )
        db.session.add(detail)
        total_amount += amount

    delivery.total_amount = total_amount

    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        raise ValueError(f"Error de integridad: {str(e)}")

    return delivery


# ============================================
# EDITAR
# ============================================
def editar(delivery_id, data):
    delivery = Delivery.query.get(delivery_id)
    if not delivery or delivery.delivery_state == "I":
        return None

    if "delivery_date" in data:
        delivery.delivery_date = datetime.fromisoformat(data["delivery_date"])

    for key, value in data.items():
        if key != "details":
            setattr(delivery, key, value)

    if "details" in data:
        incoming_details = data["details"]
        incoming_ids = [d.get("id") for d in incoming_details if d.get("id")]

        # Eliminar detalles que no vienen en incoming
        for detail in list(delivery.details):
            if detail.id not in incoming_ids:
                db.session.delete(detail)

        total_amount = 0
        for det in incoming_details:
            amount = det.get("amount", 0) or 0
            if det.get("id"):
                existing = DeliveryDetail.query.get(det["id"])
                if existing:
                    existing.medication_identifier = det["medication_identifier"]
                    existing.amount = amount
                    existing.observations = det.get("observations", "")
                    total_amount += amount
            else:
                new_detail = DeliveryDetail(
                    delivery_identifier=delivery_id,
                    medication_identifier=det["medication_identifier"],
                    amount=amount,
                    observations=det.get("observations", "")
                )
                db.session.add(new_detail)
                total_amount += amount

        delivery.total_amount = total_amount

    db.session.commit()
    return delivery


# ============================================
# ELIMINACIÓN LÓGICA
# ============================================
def eliminar(delivery_id):
    delivery = Delivery.query.get(delivery_id)
    if not delivery:
        return None
    delivery.delivery_state = "I"
    db.session.commit()
    return delivery


# ============================================
# RESTAURACIÓN
# ============================================
def restaurar(delivery_id):
    delivery = Delivery.query.get(delivery_id)
    if not delivery:
        return None
    delivery.delivery_state = "A"
    db.session.commit()
    return delivery


# ============================================
# GENERAR PDF DE DELIVERY + DETAILS
# ============================================
def generar_reporte_pdf(delivery_id):
    delivery = listar_por_id(delivery_id)
    if not delivery:
        return None

    # intentar obtener paciente (si existe)
    patient_text = f"ID: {delivery.patient_identifier}"
    if Patient:
        try:
            p = Patient.query.filter_by(id=delivery.patient_identifier).first()
            if p:
                patient_name = getattr(p, "name", None)
                patient_last_name = getattr(p, "last_name", None)
                if patient_name:
                    patient_text = f"{patient_name} {patient_last_name} (ID: {delivery.patient_identifier})"
        except Exception:
            pass

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    content = []

    # TÍTULO
    content.append(Paragraph("Reporte de Entrega", styles["Title"]))
    content.append(Spacer(1, 12))

    # CABECERA
    cabecera = [
        ["ID Entrega:", delivery.id],
        ["Paciente:", patient_text],
        ["Fecha:", str(delivery.delivery_date)],
        ["Total Cantidad:", delivery.total_amount],
        ["Notas:", delivery.notes or ""]
    ]

    table_cab = Table(cabecera, colWidths=[120, 350])
    table_cab.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))

    content.append(table_cab)
    content.append(Spacer(1, 20))

    # DETALLES
    detalles = [["ID", "Medicamento", "Cantidad", "Observaciones"]]

    for d in delivery.details:
        med = Medication.query.filter_by(id=d.medication_identifier).first()
        medication_text = f"{med.name} (ID: {d.medication_identifier})" if med else f"ID: {d.medication_identifier}"

        detalles.append([
            d.id,
            medication_text,
            d.amount,
            d.observations or ""
        ])

    tabla_det = Table(detalles, colWidths=[50, 260, 80, 160])
    tabla_det.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    content.append(Paragraph("Detalles de la Entrega", styles["Heading2"]))
    content.append(tabla_det)

    doc.build(content)

    buffer.seek(0)
    return buffer


# ============================================
# GENERAR EXCEL DE DELIVERY + DETAILS
# ============================================
def generar_reporte_excel(delivery_id):
    delivery = listar_por_id(delivery_id)
    if not delivery:
        return None

    # intentar obtener paciente (si existe)
    patient_text = f"ID: {delivery.patient_identifier}"
    if Patient:
        try:
            p = Patient.query.filter_by(id=delivery.patient_identifier).first()
            if p:
                patient_name = getattr(p, "name", None)
                patient_last_name = getattr(p, "last_name", None)
                if patient_name:
                    patient_text = f"{patient_name} {patient_last_name} (ID: {delivery.patient_identifier})"
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Entrega"

    # === CABECERA GENERAL ===
    ws.append(["ID Entrega", delivery.id])
    ws.append(["Paciente", patient_text])
    ws.append(["Fecha", str(delivery.delivery_date)])
    ws.append(["Total Cantidad", delivery.total_amount])
    ws.append(["Notas", delivery.notes or ""])
    ws.append([])

    # === CABECERA DE DETALLES ===
    header_detalle = ["ID", "Medicamento", "Cantidad", "Observaciones"]
    ws.append(header_detalle)

    # Estilo de cabecera del detalle
    header_fill = PatternFill(start_color="B7D4FF", end_color="B7D4FF", fill_type="solid")
    header_row = ws.max_row
    for col in range(1, len(header_detalle) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

    # === DETALLES ===
    for d in delivery.details:
        med = Medication.query.filter_by(id=d.medication_identifier).first()
        medication_text = f"{med.name} (ID: {d.medication_identifier})" if med else f"ID: {d.medication_identifier}"

        ws.append([
            d.id,
            medication_text,
            d.amount,
            d.observations or ""
        ])

    # === ESTILOS GENERALES ===
    for row in ws.iter_rows(min_row=1, max_col=4):
        for cell in row:
            cell.font = Font(size=12)
            # observaciones pueden necesitar alineación izquierda
            if cell.column_letter == 'D':
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # === AJUSTE AUTOMÁTICO DE COLUMNAS ===
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column].width = max_length + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
