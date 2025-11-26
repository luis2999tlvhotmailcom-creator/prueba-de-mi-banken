from app.models.purchase_detail import PurchaseDetail
from app.models.purchases import Purchase, db
from app.models.suppliers import Supplier
from app.models.medications import Medication

from sqlalchemy.exc import IntegrityError
from datetime import datetime

# PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill   
import io


# ============================================
# LISTAR SOLO ACTIVOS
# ============================================
def listar_todos():
    return Purchase.query.order_by(Purchase.id).all()


# ============================================
# LISTAR POR ID 
# ============================================
def listar_por_id(id):
    return Purchase.query.filter_by(id=id).first()


# ============================================
# CREAR
# ============================================
def crear(data):
    purchase = Purchase(**data)
    db.session.add(purchase)

    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        raise ValueError(f"Error de integridad: {str(e)}")

    return purchase


# ============================================
# EDITAR
# ============================================
def editar(purchase_id, data):
    purchase = Purchase.query.get(purchase_id)
    if not purchase:
        return None

    if "purchase_date" in data:
        purchase.purchase_date = datetime.fromisoformat(data["purchase_date"])

    for key, value in data.items():
        if key != "details":
            setattr(purchase, key, value)

    if "details" in data:
        new_details = data["details"]
        new_ids = [d.get("id") for d in new_details if d.get("id")]

        for d in purchase.details:
            if d.id not in new_ids:
                db.session.delete(d)

        for item in new_details:
            if item.get("id"):
                existing = PurchaseDetail.query.get(item["id"])
                if existing:
                    existing.amount = item["amount"]
                    existing.unit_price = item["unit_price"]
                    existing.subtotal = item["subtotal"]
                    existing.medication_identifier = item["medication_identifier"]
            else:
                new = PurchaseDetail(
                    purchase_identifier=purchase_id,
                    medication_identifier=item["medication_identifier"],
                    unit_price=item["unit_price"],
                    amount=item["amount"],
                    subtotal=item["subtotal"]
                )
                db.session.add(new)

    db.session.commit()
    return purchase


# ============================================
# ELIMINACIÓN LÓGICA
# ============================================
def eliminar(id):
    purchase = Purchase.query.get(id)
    if not purchase:
        return None

    purchase.purchase_state = "I"
    db.session.commit()
    return purchase


# ============================================
# RESTAURACIÓN
# ============================================
def restaurar(id):
    purchase = Purchase.query.get(id)
    if not purchase:
        return None

    purchase.purchase_state = "A"
    db.session.commit()
    return purchase


# ============================================
# GENERAR PDF DE PURCHASE + DETAILS
# ============================================
def generar_reporte_pdf(purchase_id):
    purchase = listar_por_id(purchase_id)
    if not purchase:
        return None

    # Obtener proveedor
    supplier = Supplier.query.filter_by(id=purchase.supplier_identifier).first()
    supplier_text = f"{supplier.company_name} (ID: {purchase.supplier_identifier})" if supplier else f"ID: {purchase.supplier_identifier}"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    content = []

    # TÍTULO
    content.append(Paragraph("Reporte de Compra", styles["Title"]))
    content.append(Spacer(1, 12))

    # CABECERA
    cabecera = [
        ["ID Compra:", purchase.id],
        ["Proveedor:", supplier_text],
        ["Fecha:", str(purchase.purchase_date)],
        ["Total Cantidad:", purchase.total_amount],
        ["Total Precio:", purchase.total_price]
    ]

    table_cab = Table(cabecera, colWidths=[120, 350])
    table_cab.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))

    content.append(table_cab)
    content.append(Spacer(1, 20))

    # DETALLES
    detalles = [["ID", "Medicamento", "Cantidad", "P. Unit", "Subtotal"]]

    for d in purchase.details:

        # Obtener medicamento
        med = Medication.query.filter_by(id=d.medication_identifier).first()
        medication_text = (
            f"{med.name} (ID: {d.medication_identifier})"
            if med else
            f"ID: {d.medication_identifier}"
        )

        detalles.append([
            d.id,
            medication_text,
            d.amount,
            d.unit_price,
            d.subtotal
        ])

    tabla_det = Table(detalles)
    tabla_det.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))

    content.append(Paragraph("Detalles de la Compra", styles["Heading2"]))
    content.append(tabla_det)

    doc.build(content)

    buffer.seek(0)
    return buffer


def generar_reporte_excel(purchase_id):
    purchase = listar_por_id(purchase_id)
    if not purchase:
        return None

    # Obtener proveedor
    supplier = Supplier.query.filter_by(id=purchase.supplier_identifier).first()
    supplier_text = (
        f"{supplier.company_name} (ID: {purchase.supplier_identifier})"
        if supplier else f"ID: {purchase.supplier_identifier}"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Compra"

    # === CABECERA GENERAL ===
    ws.append(["ID Compra", purchase.id])
    ws.append(["Proveedor", supplier_text])
    ws.append(["Fecha", str(purchase.purchase_date)])
    ws.append(["Total Cantidad", purchase.total_amount])
    ws.append(["Total Precio", purchase.total_price])

    ws.append([])  # Línea en blanco

    # === CABECERA DE DETALLES ===
    header_detalle = ["ID", "Medicamento", "Cantidad", "Precio Unit.", "Subtotal"]
    ws.append(header_detalle)

    # Estilo de cabecera del detalle
    header_fill = PatternFill(start_color="B7D4FF", end_color="B7D4FF", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=7, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

    # === DETALLES ===
    for d in purchase.details:
        med = Medication.query.filter_by(id=d.medication_identifier).first()
        medication_text = (
            f"{med.name} (ID: {d.medication_identifier})"
            if med else f"ID: {d.medication_identifier}"
        )

        ws.append([
            d.id,
            medication_text,
            d.amount,
            d.unit_price,
            d.subtotal
        ])

    # === ESTILOS GENERALES ===
    for row in ws.iter_rows(min_row=1, max_col=5):
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal="center")

    # === AJUSTE AUTOMÁTICO DE COLUMNAS ===
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass

        ws.column_dimensions[column].width = max_length + 3  # margen extra

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
